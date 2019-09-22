import openpyxl
from openpyxl.cell import cell
from codegen.config import config

def read(config):
    try:
        workbook = openpyxl.load_workbook(config.source_path, data_only=True)
        worksheet = workbook.get_sheet_by_name(config.source_tab)
        last_data_row = find_last_data_row(worksheet, config.start_search_row, config.column_range_start)
        copy_range_start = "{}{}".format(config.column_range_start, config.start_search_row)
        copy_range_end   = "{}{}".format(config.column_range_stop, last_data_row)
        return worksheet[copy_range_start:copy_range_end]
    finally:
        workbook.close()

def find_last_data_row(worksheet, start_row, column_range_start):
    column_index = cell.column_index_from_string(column_range_start)
    for current_row in range(start_row, worksheet.max_row):
        val = worksheet.cell(row = current_row, column = column_index).value
        if worksheet.cell(row = current_row, column = column_index).value == None:
            return current_row - 1
    return worksheet.max_row